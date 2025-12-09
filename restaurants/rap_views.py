from rest_framework import status, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.db.models import Count, Sum, Max
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter

from accounts.models import CustomUser
from accounts.permissions import IsRestaurantAdmin
from restaurants.models import Restaurant, Cashier, RestaurantImage
from transactions.models import Transaction
from .rap_serializers import (
    RestaurantUserSerializer,
    CashierListSerializer,
    CashierUpdateSerializer,
    RestaurantSettingsSerializer,
    RestaurantImageUploadSerializer
)


def get_admin_restaurant(user):
    """Get the restaurant associated with the admin user"""
    if hasattr(user, 'restaurant_admin_profile'):
        return user.restaurant_admin_profile.restaurant
    return None


class RestaurantUsersView(APIView):
    """
    GET /api/restaurant-admin/users/
    List users with transaction aggregations for this restaurant.
    """
    permission_classes = [IsAuthenticated, IsRestaurantAdmin]

    def get(self, request):
        restaurant = get_admin_restaurant(request.user)
        if not restaurant:
            return Response({"error": "No restaurant associated"}, status=403)

        # Get filters from query params
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        min_spent = request.query_params.get('min_spent')
        max_spent = request.query_params.get('max_spent')
        min_transactions = request.query_params.get('min_transactions')
        max_transactions = request.query_params.get('max_transactions')
        search = request.query_params.get('search')

        # Base queryset: users with transactions in this restaurant
        user_ids = Transaction.objects.filter(
            restaurant=restaurant
        ).values_list('user_id', flat=True).distinct()

        users = CustomUser.objects.filter(id__in=user_ids)

        # Search filter
        if search:
            users = users.filter(
                models.Q(phone_number__icontains=search) |
                models.Q(full_name__icontains=search)
            )

        # Annotate with aggregations
        from django.db import models as db_models
        users = users.annotate(
            total_transactions=Count(
                'transactions',
                filter=db_models.Q(transactions__restaurant=restaurant)
            ),
            total_spent_before_discount=Sum(
                'transactions__sum_before_discount',
                filter=db_models.Q(transactions__restaurant=restaurant)
            ),
            total_discount_amount=Sum(
                'transactions__discount_amount_uzs',
                filter=db_models.Q(transactions__restaurant=restaurant)
            ),
            total_spent_after_discount=Sum(
                'transactions__sum_after_discount',
                filter=db_models.Q(transactions__restaurant=restaurant)
            ),
            last_transaction_date=Max(
                'transactions__created_at',
                filter=db_models.Q(transactions__restaurant=restaurant)
            )
        )

        # Apply filters
        if date_from:
            users = users.filter(last_transaction_date__gte=date_from)
        if date_to:
            users = users.filter(last_transaction_date__lte=date_to)
        if min_spent:
            users = users.filter(total_spent_after_discount__gte=min_spent)
        if max_spent:
            users = users.filter(total_spent_after_discount__lte=max_spent)
        if min_transactions:
            users = users.filter(total_transactions__gte=min_transactions)
        if max_transactions:
            users = users.filter(total_transactions__lte=max_transactions)

        # Serialize
        data = []
        for user in users:
            data.append({
                'id': user.id,
                'full_name': user.full_name,
                'phone_number': user.phone_number,
                'total_transactions': user.total_transactions or 0,
                'total_spent_before_discount': user.total_spent_before_discount or 0,
                'total_discount_amount': user.total_discount_amount or 0,
                'total_spent_after_discount': user.total_spent_after_discount or 0,
                'last_transaction_date': user.last_transaction_date
            })

        serializer = RestaurantUserSerializer(data, many=True)
        return Response(serializer.data)


class RestaurantUsersExportView(APIView):
    """
    GET /api/restaurant-admin/users/export/
    Export users to Excel with same filters as list.
    """
    permission_classes = [IsAuthenticated, IsRestaurantAdmin]

    def get(self, request):
        restaurant = get_admin_restaurant(request.user)
        if not restaurant:
            return Response({"error": "No restaurant associated"}, status=403)

        # Same logic as list view
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        search = request.query_params.get('search')

        user_ids = Transaction.objects.filter(
            restaurant=restaurant
        ).values_list('user_id', flat=True).distinct()

        from django.db import models as db_models
        users = CustomUser.objects.filter(id__in=user_ids).annotate(
            total_transactions=Count(
                'transactions',
                filter=db_models.Q(transactions__restaurant=restaurant)
            ),
            total_spent_before_discount=Sum(
                'transactions__sum_before_discount',
                filter=db_models.Q(transactions__restaurant=restaurant)
            ),
            total_discount_amount=Sum(
                'transactions__discount_amount_uzs',
                filter=db_models.Q(transactions__restaurant=restaurant)
            ),
            total_spent_after_discount=Sum(
                'transactions__sum_after_discount',
                filter=db_models.Q(transactions__restaurant=restaurant)
            ),
            last_transaction_date=Max(
                'transactions__created_at',
                filter=db_models.Q(transactions__restaurant=restaurant)
            )
        )

        if search:
            users = users.filter(
                db_models.Q(phone_number__icontains=search) |
                db_models.Q(full_name__icontains=search)
            )
        if date_from:
            users = users.filter(last_transaction_date__gte=date_from)
        if date_to:
            users = users.filter(last_transaction_date__lte=date_to)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        # Headers
        headers = [
            'Phone Number', 'Full Name', 'Total Transactions',
            'Total Before Discount', 'Total Discount', 'Total After Discount',
            'Last Transaction'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Data rows
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1, value=user.phone_number)
            ws.cell(row=row, column=2, value=user.full_name)
            ws.cell(row=row, column=3, value=user.total_transactions or 0)
            ws.cell(row=row, column=4, value=float(user.total_spent_before_discount or 0))
            ws.cell(row=row, column=5, value=float(user.total_discount_amount or 0))
            ws.cell(row=row, column=6, value=float(user.total_spent_after_discount or 0))
            ws.cell(row=row, column=7, value=str(user.last_transaction_date) if user.last_transaction_date else '')

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{restaurant.name}_users.xlsx"'
        return response


class CashierListCreateView(APIView):
    """
    GET /api/restaurant-admin/cashiers/ - List cashiers
    POST /api/restaurant-admin/cashiers/ - Create cashier
    """
    permission_classes = [IsAuthenticated, IsRestaurantAdmin]

    def get(self, request):
        restaurant = get_admin_restaurant(request.user)
        if not restaurant:
            return Response({"error": "No restaurant associated"}, status=403)

        cashiers = Cashier.objects.filter(restaurant=restaurant)
        serializer = CashierListSerializer(cashiers, many=True)
        return Response(serializer.data)

    def post(self, request):
        restaurant = get_admin_restaurant(request.user)
        if not restaurant:
            return Response({"error": "No restaurant associated"}, status=403)

        name = request.data.get('name')
        phone_number = request.data.get('phone_number')

        if not name or not phone_number:
            return Response({"error": "name and phone_number are required"}, status=400)

        # Generate PIN
        raw_pin = Cashier.generate_pin(length=4)

        cashier = Cashier(
            restaurant=restaurant,
            name=name,
            phone_number=phone_number
        )
        cashier.set_pin(raw_pin)
        cashier.save()

        return Response({
            "cashier": CashierListSerializer(cashier).data,
            "pin_code": raw_pin,
            "message": "Save this PIN - it will not be shown again."
        }, status=201)


class CashierDetailView(APIView):
    """
    PATCH /api/restaurant-admin/cashiers/{id}/ - Update cashier
    """
    permission_classes = [IsAuthenticated, IsRestaurantAdmin]

    def patch(self, request, cashier_id):
        restaurant = get_admin_restaurant(request.user)
        if not restaurant:
            return Response({"error": "No restaurant associated"}, status=403)

        cashier = get_object_or_404(Cashier, pk=cashier_id, restaurant=restaurant)
        serializer = CashierUpdateSerializer(cashier, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(CashierListSerializer(cashier).data)


class CashierRegeneratePINView(APIView):
    """
    POST /api/restaurant-admin/cashiers/{id}/regenerate-pin/
    """
    permission_classes = [IsAuthenticated, IsRestaurantAdmin]

    def post(self, request, cashier_id):
        restaurant = get_admin_restaurant(request.user)
        if not restaurant:
            return Response({"error": "No restaurant associated"}, status=403)

        cashier = get_object_or_404(Cashier, pk=cashier_id, restaurant=restaurant)

        raw_pin = Cashier.generate_pin(length=4)
        cashier.set_pin(raw_pin)
        cashier.save()

        return Response({
            "cashier": CashierListSerializer(cashier).data,
            "pin_code": raw_pin,
            "message": "New PIN generated. Save it - it will not be shown again."
        })


class RestaurantSettingsView(APIView):
    """
    GET /api/restaurant-admin/restaurant/ - Get restaurant settings
    PATCH /api/restaurant-admin/restaurant/ - Update restaurant settings
    """
    permission_classes = [IsAuthenticated, IsRestaurantAdmin]

    def get(self, request):
        restaurant = get_admin_restaurant(request.user)
        if not restaurant:
            return Response({"error": "No restaurant associated"}, status=403)

        serializer = RestaurantSettingsSerializer(restaurant)
        return Response(serializer.data)

    def patch(self, request):
        restaurant = get_admin_restaurant(request.user)
        if not restaurant:
            return Response({"error": "No restaurant associated"}, status=403)

        serializer = RestaurantSettingsSerializer(restaurant, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(RestaurantSettingsSerializer(restaurant).data)


class RestaurantGalleryUploadView(APIView):
    """
    POST /api/restaurant-admin/restaurant/gallery/ - Upload image to gallery
    """
    permission_classes = [IsAuthenticated, IsRestaurantAdmin]

    def post(self, request):
        restaurant = get_admin_restaurant(request.user)
        if not restaurant:
            return Response({"error": "No restaurant associated"}, status=403)

        image = request.FILES.get('image')
        if not image:
            return Response({"error": "image file is required"}, status=400)

        gallery_image = RestaurantImage.objects.create(
            restaurant=restaurant,
            image=image
        )

        return Response({
            "id": str(gallery_image.id),
            "image": gallery_image.image.url
        }, status=201)
